from logweb import web_auth
from lxml import html
# import urllib
import requests
from sys import argv
# RegEx
import re
# Excel
import openpyxl
# date
from datetime import datetime


global file_name


def get_favs():
    global car_urls
    global session
    user_name, user_password = web_auth()
    # Login page
    log_url = "https://www.usedcarsni.com/login_page.php"
    # Index page
    #    req_url = "https://www.usedcarsni.com/"
    # Favorites
    fav_url = "https://www.usedcarsni.com/mystocklist.php"
    # Login data: username, password and hidden input tag
    #  'attribute name': 'attribute value'
    payload = {
        'f1': user_name,
        'f2': user_password,
        'someFrm2': '1'
    }
    with requests.Session() as session:
        session.post(log_url, data=payload)
        # Open Favorites page
        r = session.get(fav_url)
        favs = r.text
        tree = html.fromstring(favs)
        # Retrive a list of saved links from the page
        links = tree.xpath("//div[@class='car-caption hidden-md']/a/@href ")
        # Save links to List
        car_urls = []
        for e in links:
            car_urls.append(e)


def make_research():
    # global page_title
    global tree, session
#    global session
    global car_page
    #    global file_name
    #    global ref
    session = requests.Session()
    r = session.get(fav_url)
    favs = r.text
    tree = html.fromstring(favs)
    # find a hrefs that contains text 'Next'. It's a link to  a new page
    # ext = tree.xpath("//li/a[contains(text(), 'Next')]")
    if tree.xpath("//li/a[contains(text(), 'Next')]"):
        # extaact a link to the next page
        ref = tree.xpath("//li/a[contains(text(), 'Next')]/@href")
        print(fav_url)
        print(1)
        # create a list of links and add 1st link here
        ref_links = [fav_url]

        def getpagenew(ref):
            global car_page
            global tree
            #        global new_url
            # takes a link to the next page
            ref = ref[0]
            #        print(ref)
            new_url = 'http://www.usedcarsni.com' + ref
            ref_links.append(new_url)
            #        print("Link to")
            print(new_url)
            r = session.get(new_url)
            #    response = requests.get(new_url)
            car_page = r.text
            tree = html.fromstring(car_page)

        getpagenew(ref)
        num = 1
        while tree.xpath("//li/a[contains(text(), 'Next')]"):
            # numbers and print only for debug purposes
            num += 1
            print(num)
            #    print(new_url)
            ref = tree.xpath("//li/a[contains(text(), 'Next')]/@href")
            getpagenew(ref)
        else:
            pass
        # debug. checking the result
        print("Length of the List")
        print(len(ref_links))
        # print(ref_links)

        for urls in ref_links:
            print(urls)
            r = session.get(urls)
            cars = r.text
            tree = html.fromstring(cars)
            car_urls = tree.xpath("//div[@class='car-caption hidden-md']/a/@href ")
            # print(car_urls)
            retrieve_results(car_urls)
    else:
        car_urls = tree.xpath("//div[@class='car-caption hidden-md']/a/@href ")
        # print(car_urls)
        retrieve_results(car_urls)


def excel_title():
    global file_name, wb, worksheet
    file_name = 'data/' + file_name + '.xlsx'
    from openpyxl import load_workbook
    try:
        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        print(sheets)
        if worksheet in sheets:
            sheets.remove_sheet(worksheet)
        else:
            pass
#        print(now)
#        for i in sheets:
#            if now in i:
#                sheets.remove(i_sheet)
#            else:
#                pass
    # now = now + "-1"
    except:
        wb = openpyxl.Workbook()

    print("Another time")
    print(worksheet)
    wb.create_sheet(title=worksheet, index=0)
    global sheet
    sheet = wb[worksheet]
    headers = [
        'Model',
        'Year',
        'Brand',
        'Mileage',
        'Location',
        'Colour',
        'Engine Size',
        'Fuel type',
        'Gear box',
        'Doors',
        'Style',
        'Emissions',
        'Standard Tax',
        'Insurance rate',
        'Fuel consumption - Urban (mpg)',
        'Litres per KM',
        'Acceleration (0-62mph)',
        'Price',
        'Link']

    sheet.append(headers)
    wb.save(file_name)


def favorites():
    print("Get Favorites From the Account")
    get_favs()
#    print(file_name)
    print("Retrieve Results")
    retrieve_results(car_urls)


def exceldata():
    #    global sheet
    #    global j
    #    for car_spec in car_specs:
    #        i = car_specs.index(car_spec)
    #        i = i + 1
    #        cell = sheet.cell(row = j, column = i)
    #        cell.value = car_spec
    #        wb.save(file_name)
    sheet.append(car_specs)
    wb.save(file_name)


def retrieve_results(car_urls):
    # create or trunckate file with column title at the fist line
    #    global ltrkm
    global file_name, ltrkm
    # from each search page obtain individual car's links
    link_count = 0
    #   global car_urls
    for car_url in car_urls:
        link_count += 1
        car_url = car_url.replace('#Car-Tail-Url#', '')
        if 'search_type' in car_url:
            repl = re.search(r'\?[a-zA-Z0-9_=&%+]*', car_url).group()
            car_url = car_url.replace(repl, '')
        #    print(car_url)
        else:
            pass
        car_url = 'https://www.usedcarsni.com' + car_url
        print(car_url)
        # open individual car's page
        r = session.get(car_url)
        car_page = r.text
        tree = html.fromstring(car_page)
        # retrive a car's model and year
        car_name = tree.xpath("//a[@class='car-name-link']/text()")
        # strip whitespases and \n
        # copy to a new list
        car_title = []
        for car in car_name:
            car = car.rstrip()
            car_title.append(car)

        # retrive a car's price
        if tree.xpath("//span[@class='y-big-price_green y-big-price']/text()"):
            car_price = tree.xpath("//span[@class='y-big-price_green y-big-price']/text()")
        else:
            car_price = ['Sold']

        # retrive car's details such as gear, fuel, insurance rate, etc.
        #        car_tech = tree.xpath("//div[@class='technical-info']//text()")
        #        car_specs = []
        if tree.xpath("//div[@class='technical-headers'][contains(., 'Mileage')]"):
            car_mileage = tree.xpath(
                "//div[@class='technical-headers'][contains(., 'Mileage')]/following-sibling::div/text()")
        else:
            car_mileage = ['N/A']

        car_location = tree.xpath(
            "//div[@class='technical-headers'][contains(., 'Location')]/following-sibling::div/text()")
        if tree.xpath("//div[@class='technical-headers'][contains(., 'Colour')]"):
            car_colour = tree.xpath("//div[@class='technical-headers'][contains(., "
                                    "'Colour')]/following-sibling::div/text()")
        else:
            car_colour = ['N/A']

        if tree.xpath("//div[@class='technical-headers'][contains(., 'Engine Size')]"):
            car_engine = tree.xpath("//div[@class='technical-headers'][contains(., 'Engine "
                                    "Size')]/following-sibling::div/text()")
        else:
            car_engine = ['N/A']

        if tree.xpath("//div[@class='technical-headers'][contains(., 'Fuel Type')]"):
            car_fuel = tree.xpath(
                "//div[@class='technical-headers'][contains(., 'Fuel Type')]/following-sibling::div/text()")
        else:
            car_fuel = ['N/A']

        if tree.xpath("//div[@class='technical-headers'][contains(., 'Transmission')]"):
            car_trans = tree.xpath(
                "//div[@class='technical-headers'][contains(., 'Transmission')]/following-sibling::div/text()")
        else:
            car_trans = ['N/A']

        if tree.xpath("//div[@class='technical-headers'][contains(., 'Doors')]"):
            car_doors = tree.xpath(
                "//div[@class='technical-headers'][contains(., 'Doors')]/following-sibling::div/text()")
        else:
            car_doors = ['N/A']

        if tree.xpath("//div[@class='technical-headers'][contains(., 'Body Style')]"):
            car_body = tree.xpath("//div[@class='technical-headers'][contains(., 'Body "
                                  "Style')]/following-sibling::div/text()")
        else:
            car_body = ['N/A']

        if tree.xpath("//div[@class='technical-headers'][contains(., 'CO2 Emission')]"):
            car_co2 = tree.xpath(
                "//div[@class='technical-headers'][contains(., 'CO2 Emission')]/following-sibling::div/text()")
        else:
            car_co2 = ['N/A']

        if tree.xpath("//div[@class='technical-headers'][contains(., 'Standard Tax')]"):
            car_t = tree.xpath(
                "//div[@class='technical-headers'][contains(., 'Standard Tax')]/following-sibling::div/a/text()")
            car_tax = []
            for x in car_t:
                x = x.strip()
                car_tax.append(x)
        else:
            car_tax = ['N/A']

        if tree.xpath("//div[@class='technical-headers'][contains(., 'Insurance')]"):
            car_ins = tree.xpath(
                "//div[@class='technical-headers'][contains(., 'Insurance')]/following-sibling::div/a[1]/text()")
            car_insurance = []
            for x in car_ins:
                x = x.strip()
                car_insurance.append(x)
                if not car_insurance:
                    car_insurance = ['No Data']
        else:
            car_insurance = ['N/A']

        # retrive Fuel Consumption - Urban
        if tree.xpath("//td[@role='rowheader'][contains(text(), 'Fuel Consumption - Urban')]"):
            urban = tree.xpath(
                "//td[@role='rowheader'][contains(text(), 'Fuel Consumption - Urban')]/following-sibling::td/text()")
            urban_mpg = []
            ltrkm = ''
            for elem in urban:
                elem = elem.strip()
                mls = elem.replace(' mpg', '')
                mls = float(mls)
                ltr = 282.48 / mls
                ltr = round(ltr)
                ltr = int(ltr)
                ltr = str(ltr) + ' l/km'
                ltrkm = [ltr]
                assert isinstance(elem, object)
                urban_mpg.append(elem)
        else:
            urban_mpg = ['N/A']
            ltrkm = ['N/A']
        # clean up data and copy to a new list
        if tree.xpath("//td[@role='rowheader'][contains(text(), 'Acceleration (0-62mph)')]"):
            acceleration = tree.xpath(
                "//td[@role='rowheader'][contains(text(), 'Acceleration (0-62mph)')]/following-sibling::td/text()")
            car_acceleration = []
            for elem in acceleration:
                elem = elem.strip()
                car_acceleration.append(elem)
        else:
            car_acceleration = ['N/A']

        #        car_year = re.search('20([1][0-9]||[0][0-9]||[20])', car_title[0]).group()
        if re.search(r'20[0-1][0-9]|[2][0]|19[9][0-9]', car_title[0]):
            car_year = re.search(r'20[0-1][0-9]|[2][0]|19[9][0-9]', car_title[0]).group()
            car_year = [car_year]
        else:
            car_year = ['No Data']
        #    print(car_year[0])

        cars = tree.xpath("//select[@id='make']/option/text()")
        for brand in cars:
            #    for brand in brands:
            car_brand = re.search(brand, car)
            if car_brand:
                car_brand = car_brand.group()
                car_name = [car_brand]
            #            print(car_brand[0])
            else:
                pass
        global car_specs
        # merge data to the one list
        car_specs = [*car_title,
                     *car_year,
                     *car_name,
                     *car_mileage,
                     *car_location,
                     *car_colour,
                     *car_engine,
                     *car_fuel,
                     *car_trans,
                     *car_doors,
                     *car_body,
                     *car_co2,
                     *car_tax,
                     *car_insurance,
                     *urban_mpg,
                     *ltrkm,
                     *car_acceleration,
                     *car_price]
        car_specs.append(car_url)
        #        print(car_specs)
        #        global j
        #        j = j + 1
        #       csvdata()
        exceldata()
    print(f'There are {link_count} cars.')


if __name__ == "__main__":
    if len(argv) > 1:
        # read req.textif req.text exist and not empty
        try:
            file = open('req.txt')
        except IOError as e:
            print("You don't have req.txt file. Reading your favorites list from your Account")
            favorites()
        else:
            with file:
                print("Read the file")
                fav_url = file.read()
                script, cmdarg = argv
                worksheet = cmdarg
                file_name = "usedcarsni"
                # Function start reseach
                # table_title()
                excel_title()
                print("Make a research")
                make_research()
                if not fav_url:
                    print("There is no url in req.txt file")
                    pass
    # That has a Function to read list
    else:
        file_name = 'favorites'
        now = datetime.now().date()
        worksheet = str(now)
        #        print(now)
        # table_title()
        excel_title()
        favorites()
