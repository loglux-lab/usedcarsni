import requests
import re
import csv
import json
import openpyxl
from lxml import html
import pandas as pd
from storage import Storage
from storage import Operations
from datetime import date
import sqlite3

class Cars:
    def __init__(self, url, file_name="cars"):
        """ TODO: Create a method for dealing with a list of favorites
            TODO: Create save_to_json method.
        """
        self.url = url
        self.page_url = url
        self.file_name = "cars"
        self.base_url = "https://www.usedcarsni.com"
        self.page_number = "&pagepc0="
        self.car_catalogue = []
        self.car_columns = []
        self.result = False
        self.session = requests.Session()
        self.current_date = date.today()
        self.connect()
        self.car_values = ''
        self.insert_car = """INSERT INTO {}  {}  VALUES ( {} );"""
        self.all_del = """DELETE FROM {} WHERE id IN ({})"""
        self.select_table = """SELECT * from {} """
        self.update_car = """UPDATE cars SET
        Price = ? WHERE Id IN ({});"""
        self.insert_date = """INSERT INTO price_watch (
        H_Date,
        H_Price,
        Id)
        VALUES (?, ?, ?);
         """
        self.price_select = """SELECT DISTINCT H_Date, H_Price, Id 
        FROM price_watch WHERE Id= ?
        AND 
        date(H_date) = (SELECT MAX(date(H_Date)) FROM price_watch)"""
        self.row_check = """SELECT Price from cars WHERE id IN ({})"""


    def start(self):
        page_scope = self.check
        self.directory(page_scope)

    def connect(self):
        self.session.headers.update({'User-Agent': 'Mozilla / 5.0 (Linux; Android  6.0)'})
        self.r = self.session.get(self.url)
        self.tree = html.fromstring(self.r.text)
        return self.tree

    @property
    def check(self):
        page_scope = self.tree.xpath("//div[@class='page-control-label']")[0].text
        page_scope = page_scope.strip()
        page_scope = re.sub(r'[0-9]*\sto\s[0-9]*\sof\s', '', page_scope)
        page_total = int(page_scope)
        page_scope = page_total // 20 + 1
        if page_total % 20 > 1:
            page_scope = page_scope + 1
        return page_scope

    def directory(self, page_scope):
         for x in range(1, page_scope):
            x = str(x)
            self.level_page = self.url + self.page_number + x
            self.tree = Cars(self.level_page).connect()
            print(x + " " + self.level_page)
            self.car_pages()

    def car_pages(self):
        self.car_urls = self.tree.xpath("//div[@class='car-caption hidden-md']/a/@href ")
        for self.car_url in self.car_urls:
            """ removing  #Car-Tail-Url# tag from the end of a link """
            self.car_url = self.car_url.replace('#Car-Tail-Url#', '')
            """ removing 'search_type' tag from the link """
            if 'search_type' in self.car_url:
               self.repl = re.search(r'\?[a-zA-Z0-9_=&%+]*', self.car_url).group()
               """ Retreive information by following this link. """
               self.page_url = self.base_url + self.car_url.replace(self.repl, '')
               self.tree = Cars(self.page_url).connect()
               print(self.page_url)
               self.parser()
        #return self.car_url

    def parser(self):
        """ Working with an individual page """
        self.car_description = {}
        self.car_title = []
        self.car_name = self.tree.xpath("//a[@class='car-name-link']/text()")
        [self.car_title.append(car.rstrip().replace('  ', ' ')) for car in self.car_name]
        if self.tree.xpath("//span[@class='y-big-price_green y-big-price']/text()"):
            self.car_price = self.tree.xpath("//span[@class='y-big-price_green y-big-price']/text()")
        else:
            self.car_price = ['Sold']
        if self.tree.xpath("//div[@class='technical-headers'][contains(., 'Mileage')]"):
            self.car_mileage = self.tree.xpath(
                "//div[@class='technical-headers'][contains(., 'Mileage')]/following-sibling::div/text()")
        else:
            self.car_mileage = ['N/A']
        if self.tree.xpath("//div[@class='technical-headers'][contains(., 'Location')]/following-sibling::div/text()"):
            self.car_location = self.tree.xpath(
            "//div[@class='technical-headers'][contains(., 'Location')]/following-sibling::div/text()")
        else:
            self.car_location = ['N/A']
        if self.tree.xpath("//div[@class='technical-headers'][contains(., 'Colour')]"):
            self.car_colour = self.tree.xpath("//div[@class='technical-headers'][contains(., "
                                    "'Colour')]/following-sibling::div/text()")
        else:
            self.car_colour = ['N/A']
        if self.tree.xpath("//div[@class='technical-headers'][contains(., 'Engine Size')]"):
            self.car_engine = self.tree.xpath("//div[@class='technical-headers'][contains(., 'Engine "
                                    "Size')]/following-sibling::div/text()")
        else:
            self.car_engine = ['N/A']
        if self.tree.xpath("//div[@class='technical-headers'][contains(., 'Fuel Type')]"):
            self.car_fuel = self.tree.xpath(
                "//div[@class='technical-headers'][contains(., 'Fuel Type')]/following-sibling::div/text()")
        else:
            self.car_fuel = ['N/A']
        if self.tree.xpath("//div[@class='technical-headers'][contains(., 'Transmission')]"):
            self.car_trans = self.tree.xpath(
                "//div[@class='technical-headers'][contains(., 'Transmission')]/following-sibling::div/text()")
        else:
            self.car_trans = ['N/A']
        if self.tree.xpath("//div[@class='technical-headers'][contains(., 'Doors')]"):
            self.car_doors = self.tree.xpath(
                "//div[@class='technical-headers'][contains(., 'Doors')]/following-sibling::div/text()")
        else:
            self.car_doors = ['N/A']
        if self.tree.xpath("//div[@class='technical-headers'][contains(., 'Body Style')]"):
            self.car_body = self.tree.xpath("//div[@class='technical-headers'][contains(., 'Body "
                                  "Style')]/following-sibling::div/text()")
        else:
            self.car_body = ['N/A']
        if self.tree.xpath("//div[@class='technical-headers'][contains(., 'CO2 Emission')]"):
            self.car_co2 = self.tree.xpath(
                "//div[@class='technical-headers'][contains(., 'CO2 Emission')]/following-sibling::div/text()")
        else:
            self.car_co2 = ['N/A']

        if self.tree.xpath("//div[@class='technical-headers'][contains(., 'Standard Tax')]"):
            self.car_t = self.tree.xpath(
                "//div[@class='technical-headers'][contains(., 'Standard Tax')]/following-sibling::div/a/text()")
            self.car_tax = []
            [ self.car_tax.append(x.strip()) for x in self.car_t ]
            if self.car_tax == []:
                self.car_tax = ['Foo']
        else:
            self.car_tax = ['N/A']
        if self.tree.xpath("//div[@class='technical-headers'][contains(., 'Insurance')]"):
            self.car_ins = self.tree.xpath(
                "//div[@class='technical-headers'][contains(., 'Insurance')]/following-sibling::div/a[1]/text()")
            self.car_insurance = []
            [ self.car_insurance.append(x.strip()) for x in self.car_ins ]
            if not self.car_insurance:
                self.car_insurance = ['No Data']
        else:
            self.car_insurance = ['N/A']
        if self.tree.xpath("//td[@role='rowheader'][contains(text(), 'Fuel Consumption - Urban')]"):
            self.urban_mpg = self.tree.xpath(
                "//td[@role='rowheader'][contains(text(), 'Fuel Consumption - Urban')]/following-sibling::td/text()")
            self.ltrkm = ''
            self.mpg = []
            [ self.mpg.append(elem.strip().replace(' mpg', '')) for elem in self.urban_mpg ]
            [self.car_title.append(car.rstrip().replace('  ', ' ')) for car in self.car_name]
            self.urban_mpg = self.mpg[0] + ' mpg'
            self.mpg = float(self.mpg[0])
            if self.mpg > 0:
                self.ltrkm = str(int(round((282.48 / self.mpg)))) + ' l/100km'
            else:
                self.ltrkm = 'N/A'
                self.urban_mpg = 'N/A'
        if self.tree.xpath("//td[@role='rowheader'][contains(text(), 'Acceleration (0-62mph)')]"):
            self.acceleration = self.tree.xpath(
                "//td[@role='rowheader'][contains(text(), 'Acceleration (0-62mph)')]/following-sibling::td/text()")
            self.car_acceleration = []
            [ self.car_acceleration.append(elem.strip()) for elem in self.acceleration ]
        else:
            self.car_acceleration = ['N/A']
        if self.tree.xpath("//td[@role='rowheader'][contains(text(), 'Trim')]"):
            self.trim = self.tree.xpath(
                "//td[@role='rowheader'][contains(text(), 'Trim')]/following-sibling::td/text()")
            self.car_trim = []
            [ self.car_trim.append(elem.strip()) for elem in self.trim ]
        else:
            self.car_trim = ['N/A']
        self.url_info = self.page_url.replace(self.base_url + '/', '').split('-')
        self.car_year = self.url_info[0]
        self.car_make = self.url_info[1]
        self.car_model = self.url_info[2]
        self.car_id = self.url_info[-1]
        """ Creating a dictionary """
        """ TODO: a separete method or even separate class??? """
        #self.car_description['Name'] = self.car_title[0]
        self.car_description['Make'] = self.car_make
        self.car_description['Model'] = self.car_model
        self.car_description['Trim'] = self.car_trim[0]
        self.car_description['Year'] = self.car_year
        self.car_description['Price'] = self.car_price[0]
        self.car_description['Mileage'] = self.car_mileage[0]
        #self.car_description['Location'] = self.car_location[0]
        #self.car_description['Colour'] = self.car_colour[0]
        self.car_description['Engine'] = self.car_engine[0]
        self.car_description['Fuel'] = self.car_fuel[0]
        self.car_description['Transmission'] = self.car_trans[0]
        #self.car_description['Doors'] = self.car_doors[0]
        #self.car_description['Body'] = self.car_body[0]
        #self.car_description['CO2'] = self.car_co2[0]
        self.car_description['Tax'] = self.car_tax[0]
        self.car_description['Insurance'] = self.car_insurance[0]
        self.car_description['MPG'] = self.urban_mpg
        self.car_description['KM'] = self.ltrkm
        self.car_description['Acceleration'] = self.car_acceleration[0]
        self.car_description['Link'] = self.page_url
        self.car_description['Id'] = int(self.car_id)
        """ Adding a dictionary to a big list (declared in __init__) """
        self.car_catalogue.append(self.car_description)

        """ Creating a title for saving in csv/excel tables"""
        self.car_columns = list(self.car_catalogue[0].keys())
        print(self.car_description)

    def results(self):
        print("Total in the list: " + str(len(self.car_catalogue)))
        self.result = True

    def save_to_json(self):
        """ TODO: Save a list of dictionaries to json.file, if it possible"""
        with open(self.file_name + '.json', 'w') as fp:
            [ json.dump(data, fp) for data in self.car_catalogue ]

            #json.dump(self.car_catalogue, fp)

    def save_to_csv(self):
        """ TODO: create an exception if file is not exists or busy """
        csv_file = self.file_name + ".csv"
        try:
            with open(csv_file, 'w') as csvfile:
                self.writer = csv.DictWriter(csvfile, fieldnames=self.car_columns)
                self.writer.writeheader()
                [ self.writer.writerow(data) for data in self.car_catalogue ]
                #for data in self.car_catalogue:
                 #   self.writer.writerow(data)
        except IOError:
            print("I/O error")

    def save_to_excel(self):
        """ TODO: Revise this function save_to_excel """
        from openpyxl import load_workbook
        sheet_date = str(self.current_date)
        full_file_name = self.file_name + '.xlsx'
        sheet_name = sheet_date + " - " + str(self.car_catalogue[0]['Make'])  + " - " + str(self.car_catalogue[0]['Model'])
        try:
            wb = load_workbook(full_file_name)
            sheets = wb.sheetnames  # Returns a worksheets by its names
            print(f"The file has sheets: {sheets}") # Printing excisting sheets
            for w in sheets:  # check if we have worksheet in the list
                if w == sheet_name or w == 'Sheet':
                    wb.remove(wb[w])  # Remove worksheet from this workbook
                    wb.save(full_file_name)  # Save the results
            worksheet = sheet_name
            wb.create_sheet(title=worksheet, index=0)  # Creating a worksheet
            sheet = wb[worksheet] # Opening a worksheet
            headers = self.car_columns  #
            sheet.append(headers)
            wb.save(full_file_name)
            [ sheet.append(list(data.values())) for data in self.car_catalogue ]
            wb.save(full_file_name)
        except KeyError as e:
            print("Worksheet {worksheet} does not exist.".format(worksheet=worksheet))

    def save_to_excel2(self):
        """ TODO: Revise this function save_to_excel """
        from openpyxl import load_workbook
        full_file_name = self.file_name + '.xlsx'
        try:
            wb = load_workbook(full_file_name)
            sheets = wb.sheetnames  # Returns a worksheets by its names
            print(f"The file has sheets: {sheets}")
            for worksheet in sheets:  # check if we have worksheet in the list
                if worksheet == self.file_name or worksheet == 'Sheet':
                    wb.remove(wb[worksheet])  # Remove worksheet from this workbook
                    wb.save(full_file_name)  # Save the results
            else:
                print("The first time today?")
                pass
        except:
            worksheet = self.file_name
            print(worksheet)
            wb = openpyxl.Workbook()
            wb.create_sheet(title=worksheet, index=0)
            try:
                sheet = wb[worksheet]
                headers = self.car_columns  #
                sheet.append(headers)
                wb.save(full_file_name)
                [ sheet.append(list(data.values())) for data in self.car_catalogue ]
                wb.save(full_file_name)
            except KeyError as e:
                print("Worksheet {worksheet} does not exist.".format(worksheet=worksheet))

    def print_table(self):
        """ Pretty print a list of dictionaries (myDict) as a dynamically sized table.
        If column names (colList) aren't specified, they will show in random order.
        Author: Thierry Husson - Use it as you want but don't blame me.
        """
        # https://stackoverflow.com/questions/17330139/python-printing-a-dictionary-as-a-horizontal-table-with-headers
        car_catalogue = []
        for d in self.car_catalogue:
            del d['Name']
            car_catalogue.append(d)
        del self.car_columns[0]
        my_list = [self.car_columns]  # 1st row = header
        for item in car_catalogue:
            my_list.append([str(item[col] if item[col] is not None else '') for col in self.car_columns])
        col_size = [max(map(len, col)) for col in zip(*my_list)]
        format_str = ' | '.join(["{{:<{}}}".format(i) for i in col_size])
        my_list.insert(1, ['-' * i for i in col_size])  # Seperating line
        for item in my_list:
            print(format_str.format(*item))

    def input_url(self):
        """ TODO: Input Option. File name, link, parameters"""
        self.url = input("Search URL: ")
        print(self.url)
        #self.connect()
        return  self.url

    def pd_table(self):
        #columns = list(catalogue[0].keys())
        df = pd.DataFrame(self.car_catalogue)
        #df[['Name', 'Qualification']]
        df = df[['Make', 'Model', 'Trim', 'Year', 'Price', 'Mileage', 'Engine', 'Fuel',
                'Transmission', 'Tax', 'Insurance', 'MPG', 'KM', 'Acceleration', 'Link']]
        pd.set_option('max_columns', 20)
        #pd.set_option('display.max_rows', None)
        #pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        df = df.set_index("Make")
        #pd.set_option('display.max_colwidth', -1)
        #pd.set_option('display.max_columns', 999)
        # all rows:
        # pandas.options.display.max_rows
        print(df)

    def save_to_db(self):
        fresh = []
        for x in self.car_catalogue:
            fresh.append(x['Id'].replace(',', ''))
        sql_id = """ SELECT Id FROM cars """
        sql_diff = """ SELECT * FROM cars WHERE Id = ?"""
        with Storage() as cursor:
            cursor.execute(sql_id)
            old = cursor.fetchall()
            # Res = [x for x in Ans if x in Word]
            # https://ru.stackoverflow.com/questions/427942/%D0%A1%D1%80%D0%B0%D0%B2%D0%BD%D0%B5%D0%BD%D0%B8%D0%B5-2-%D1%83%D1%85-%D1%81%D0%BF%D0%B8%D1%81%D0%BA%D0%BE%D0%B2-%D0%B2-python/427949
            old_one = [ item for t in old for item in t]
            fresh_one = []
            for i in fresh:
                i = int(i)
                fresh_one.append(i)
            id_diff = list(set(old_one) - set(fresh_one))
            inv_diff = list(set(fresh_one) - set(old_one))
            print(id_diff)
            print("Inversion: " + str(inv_diff))
            cursor.execute(sql_diff, id_diff)
            dropped_cars = cursor.fetchall()
            print(dropped_cars)

    def db_operations(self):
        """ TODO: Move all queries to __init__ or even create a new calss or move everythng to the class Storage """
        """ TODO: Create prices monitoring """
        self.car_values = len(self.car_columns) * "?, "
        self.car_values = self.car_values.rstrip(', ')
        """ Creating a list of lists of values """
        cars = []
        [cars.append(tuple(car.values())) for car in self.car_catalogue]
        with Storage() as cursor:
            try:
                """ Open cars """
                cursor.execute(self.select_table.format("cars"))
                previous_data = cursor.fetchall()
                """ if it doesn't exist """
            except sqlite3.OperationalError:
                """ Create a new tables: cars and price_watch """
                Operations().create_tables()
                """ Insert Data into cars """
                cursor.executemany(self.insert_car.format("cars", tuple(self.car_columns), self.car_values), cars)
            else:
                """ Check New, Deleted and Prices """
                """ list( set(A) - set(B) ) and vice versa """
                fresh_id = []
                [ fresh_id.append(car['Id']) for car in self.car_catalogue ]
                print("fresh id: " + str(fresh_id))
                previous_id = []
                [ previous_id.append(car[15]) for car in previous_data ]
                print("previous id: " + str(previous_id))
                """ Deleted """
                removed_id = (list(set(previous_id) - set(fresh_id)))
                print("removed id : " + str(removed_id))
                if removed_id:
                    """ removing deleted """
                    del_cars = []
                    [del_cars.append(car) for car in previous_data if car[15] in removed_id ]
                    cursor.executemany(self.insert_car.format("old", tuple(self.car_columns), self.car_values), del_cars)
                    """ delete removed cars: """
                    car_rm = "?, " * len(removed_id)
                    car_rm = car_rm.rstrip(', ')
                    cursor.execute(self.all_del.format("cars", car_rm), removed_id)
                """ New """
                new_id = (list( set(fresh_id) - set(previous_id)) )
                print("New id: " + str(new_id))
                if new_id:
                    new_cars = []
                    [ new_cars.append(car) for car in cars if car[15] in new_id ]
                    print(new_cars)
                    print("New id: " + str(new_id))
                    cursor.executemany(self.insert_car.format("cars", tuple(self.car_columns), self.car_values), new_cars)
                """ Checking prices here ? """
                car_prices = []
                [ car_prices.append(car) for car in self.car_catalogue if car['Id'] not in removed_id or new_id ]
                """ Actual data: """
                car_id = (tuple( set(previous_id) - set(removed_id)))
                car_data = "?, " * len(car_id)
                car_data = car_data.rstrip(', ')
                cursor.execute(self.row_check.format(car_data) , car_id)
                #(car_prices)
                """ Comparison here """


if __name__ == '__main__':
    all_renault2016 = "https://www.usedcarsni.com/search_results.php?keywords=&make=24&model=1170&fuel_type=0&trans_type=0&age_from=2016&age_to=0&price_from=0&price_to=0&user_type=0&mileage_to=0&body_style=12&doors%5B%5D=5&keywords=&distance_enabled=0&distance_postcode=&homepage_search_attr=1&tab_id=0&search_type=1"
    #search_url = str(input("Search URL: "))
    #file_name = str(input("File name: "))
    #if not search_url and not file_name:
    search_url = "https://www.usedcarsni.com/search_results.php?search_type=1&make=24&fuel_type=2&age_from=2016&price_from=0&user_type=2%7C4&model=1170&trans_type=0&age_to=0&price_to=0&mileage_to=0&keywords=&distance_enabled=1&distance_postcode=&body_style=12&doors%5B%5D=5"
    hyundai10 = "https://www.usedcarsni.com/search_results.php?search_type=1&make=9&fuel_type=0&age_from=0&price_from=0&user_type=0&model=17036939&trans_type=0&age_to=0&price_to=0&mileage_to=0&keywords=&distance_enabled=1&distance_postcode=&body_style=0"
    hyundaii20 = "https://www.usedcarsni.com/search_results.php?search_type=1&make=9&fuel_type=0&age_from=0&price_from=0&user_type=2%7C4&model=17375267&trans_type=0&age_to=0&price_to=0&mileage_to=0&keywords=&distance_enabled=1&distance_postcode=&body_style=0"
    motor = Cars(search_url)
    #motor.connect()
    #motor.check
    #motor.directory()
    motor.start()
    motor.results()
    motor.db_operations()
    motor.pd_table()
    motor.save_to_excel()


"""    motor.save_to_csv()
    motor.save_to_json()
    motor.save_to_excel()
    #motor.print_table()
    motor.pd_table()
"""